import os
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from flask import Flask, render_template, request
from flask import *  
from flask_mail import *  


app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = "flask proj/upload" 

app.config["MAIL_SERVER"]='smtp.gmail.com'  
app.config["MAIL_PORT"] = 465  
app.config["MAIL_USERNAME"] = 'ABC@gmail.com'  
app.config['MAIL_PASSWORD'] = '******'  
app.config['MAIL_USE_TLS'] = False  
app.config['MAIL_USE_SSL'] = True  
mail = Mail(app) 

def send_email(path2):
    student_dict={}
    f = open(path2, "r")
    reader = csv.reader(f)
    for row in reader:
        if row[6] == "Roll Number":
            continue
        else:
            student_dict[row[6]] = row[4]
    for i in student_dict:
        msg = Message(subject = "Marksheet", sender = "ABC@gmail.com", recipients=[student_dict[i]])
        with app.open_resource("./sample_output/marksheet/") as fp:
            msg.attach("{}.xlsx".format(i),fp.read()) 
            mail.send(msg)     


def concise(path2, positive, negative):
    if not os.path.isdir("sample_output"):
        os.path.mkdir("sample_output")
    if not os.path.isdir(os.path.join("sample_output", "coincise_marksheet")):
        os.path.join("sample_output", "coincise_marksheet")

    f = open(path2, "r")
    reader = csv.reader(f)
    dict1 = {}
    for row in reader:
        if row[0] == "Timestamp":
            continue
        elif row[6] == "ANSWER":
            temp = 1
            x = range(7,10000)
            for n in x:
                try:
                    dict1[n-6] = row[n] #dict store correct ans key
                except:
                    break

    f.close()

    f = open(path2, "r")
    reader = csv.reader(f)
    x = 1
    wb = Workbook()
    sheet = wb.active

    for row in reader:
        dict2 = {}
        right = 0
        wrong = 0
        na = 0
      
        z = range(6)
        for n in z:
            sheet.cell(row = x, column = n+1).value = row[n]
        z = range(7,10000)
        for n in z:
            try:
                sheet.cell(row = x, column = n+1).value = row[n-1]
            except:
                y = n
                break
        for n in z:
            try:
                dict2[n-6] = row[n]
            except:
                break

        for n in z:
            try:
                if dict1[n-6] == dict2[n-6]:
                    right += 1
                elif dict2[n-6] == "":
                    na += 1
                else:
                    wrong += 1

            except:
                break
      
        total = int(int(len(dict1)) * int(positive))

        if row[0] == "Timestamp":
            sheet['G1'] = "Score_After_Negative"
            sheet.cell(row = x, column = y+1).value = "statusAns"
        else:
            sheet.cell(row = x, column= 7).value = str((int(positive) *int(right)) + (int(negative) * int(wrong))) + "/" + str(int(total))
            sheet.cell(row = x, column = y+1).value = "[" + str(right) + ", " + str(wrong) + ", " + str(na) + "]"
     
        x += 1

    wb.save(os.path.join(os.getcwd(), "sample_output",
                    "coincise_marksheet.xlsx"))
    f.close()


def generate_marksheet(path2, positive, negative):
    if not os.path.isdir("sample_output"):
        os.mkdir("sample_output")
    if not os.path.isdir(os.path.join("sample_output", "marksheet")):
        os.mkdir(os.path.join("sample_output", "marksheet"))

    f = open(path2, "r")
    reader = csv.reader(f)

    temp = 0
    dict = {}

    for row in reader:
        if row[0] == "Timestamp":
            continue
        elif row[6] == "ANSWER":
            temp = 1
            x = range(7, 35)
            for n in x:
                try:
                    dict[n-6] = row[n]  # dict store correct ans key
                except:
                    break

    if temp == 0:
        print("Wrong File")
    else:
        f = open(path2, "r")
        reader = csv.reader(f)
        for row in reader:
            if row[0] == "Timestamp":
                continue
            Right = 0
            Wrong = 0
            Not_Attempted = 0
            wb = Workbook()
            wb.create_sheet(index=0, title='quiz')
            sheet = wb['quiz']

            sheet.column_dimensions['A'].width = 17.09
            sheet.column_dimensions['B'].width = 17.09
            sheet.column_dimensions['C'].width = 17.09
            sheet.column_dimensions['D'].width = 17.09
            sheet.column_dimensions['E'].width = 17.09

            sheet.row_dimensions[5].height = 22.50

            sheet.merge_cells('A5:E5')
            sheet['A5'].font = Font(
                name='Century', size=18, bold=True, underline='single')
            sheet['A5'].alignment = Alignment(
                horizontal='center', vertical='bottom')
            sheet['A5'] = 'Mark Sheet'

            img = openpyxl.drawing.image.Image('IITP_Logo.png')
            img.anchor = 'A1'
            sheet.add_image(img)

            sheet['A6'] = "Name:"
            sheet['A6'].font = Font(name='Century', size=12)
            sheet['A6'].alignment = Alignment(
                horizontal='right', vertical='bottom')

            sheet['D6'] = "Exam:"
            sheet['D6'].font = Font(name='Century', size=12)
            sheet['D6'].alignment = Alignment(
                horizontal='right', vertical='bottom')

            sheet['A7'] = "Roll Number:"
            sheet['A7'].font = Font(name='Century', size=12)
            sheet['A7'].alignment = Alignment(
                horizontal='right', vertical='bottom')

            sheet['E6'] = "quiz"
            sheet['E6'].font = Font(name='Century', size=12, bold=True)
            sheet['E6'].alignment = Alignment(
                horizontal='left', vertical='bottom')

            sheet['B6'] = row[3]
            sheet['B6'].font = Font(name='Century', size=12, bold=True)
            sheet['B6'].alignment = Alignment(
                horizontal='left', vertical='bottom')

            sheet['B7'] = row[6]
            sheet['B7'].font = Font(name='Century', size=12, bold=True)
            sheet['B7'].alignment = Alignment(
                horizontal='left', vertical='bottom')

            sheet['B9'] = "Right"
            sheet['B9'].font = Font(name='Century', size=12, bold=True)
            sheet['B9'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['C9'] = "Wrong"
            sheet['C9'].font = Font(name='Century', size=12, bold=True)
            sheet['C9'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['D9'] = "Not Attempt"
            sheet['D9'].font = Font(name='Century', size=12, bold=True)
            sheet['D9'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['E9'] = "Max"
            sheet['E9'].font = Font(name='Century', size=12, bold=True)
            sheet['E9'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['A10'] = "No."
            sheet['A10'].font = Font(name='Century', size=12, bold=True)
            sheet['A10'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['A11'] = "Marking"
            sheet['A11'].font = Font(name='Century', size=12, bold=True)
            sheet['A11'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['A12'] = "Total"
            sheet['A12'].font = Font(name='Century', size=12, bold=True)
            sheet['A12'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['A15'].border = Border(right=Side(border_style='thin', color='00000000'),
                                         top=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'))
            sheet['A15'] = "Student Ans"
            sheet['A15'].font = Font(name='Century', size=12, bold=True)
            sheet['A15'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['B15'] = "Correct Ans"
            sheet['B15'].font = openpyxl.styles.Font(
                name='Century', size=12, bold=True)
            sheet['B15'].alignment = openpyxl.styles.Alignment(
                horizontal='center', vertical='bottom')
            sheet['B15'].border = Border(right=Side(border_style='thin', color='00000000'),
                                         top=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'))

            sheet['B11'] = positive
            sheet['B11'].font = Font(name='Century', size=12, color='008000')
            sheet['B11'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['C11'] = negative
            sheet['C11'].font = Font(name='Century', size=12, color='FF0000')
            sheet['C11'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['D11'] = "0"
            sheet['D11'].font = Font(name='Century', size=12)
            sheet['D11'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            for i in range(9, 13):
                for j in range(1, 6):
                    sheet.cell(row=i, column=j).border = Border(right=Side(border_style='thin', color='00000000'),
                                                                top=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'))

            z = range(len(dict))

            for n in z:
                sheet.cell(row=16 + n, column=2).value = dict[n+1]
                sheet.cell(row=16 + n, column=2).font = Font(name='Century',
                                                             size=12, bold=False, underline='none', color='0000FF')
                sheet.cell(
                    row=16 + n, column=2).alignment = Alignment(horizontal='center', vertical='bottom')
                sheet.cell(row=16 + n, column=2).border = Border(right=Side(border_style='thin', color='00000000'),
                                                                 top=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'))

                if row[n+7] == dict[n+1]:
                    sheet.cell(row=16 + n, column=1).value = row[n+7]
                    sheet.cell(row=16 + n, column=1).font = Font(name='Century',
                                                                 size=12, bold=False, underline='none', color='008000')
                    sheet.cell(
                        row=16 + n, column=1).alignment = Alignment(horizontal='center', vertical='bottom')
                    sheet.cell(row=16 + n, column=1).border = Border(right=Side(border_style='thin', color='00000000'),
                                                                     top=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'))
                    Right += 1

                elif row[n+7] == "":
                    sheet.cell(row=16 + n, column=1).border = Border(right=Side(border_style='thin', color='00000000'),
                                                                     top=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'))
                    Not_Attempted += 1

                else:
                    sheet.cell(row=16 + n, column=1).value = row[n+7]
                    sheet.cell(row=16 + n, column=1).font = Font(name='Century',
                                                                 size=12, bold=False, underline='none', color='FF0000')
                    sheet.cell(
                        row=16 + n, column=1).alignment = Alignment(horizontal='center', vertical='bottom')
                    sheet.cell(row=16 + n, column=1).border = Border(right=Side(border_style='thin', color='00000000'),
                                                                     top=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'))
                    Wrong += 1

            sheet['B10'] = Right
            sheet['B10'].font = Font(name='Century', size=12, color='008000')
            sheet['B10'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['E10'] = 28
            sheet['E10'].font = Font(name='Century', size=12)
            sheet['E10'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['C10'] = Wrong
            sheet['C10'].font = Font(name='Century', size=12, color='FF0000')
            sheet['C10'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            sheet['D10'] = Not_Attempted
            sheet['D10'].font = Font(name='Century', size=12)
            sheet['D10'].alignment = openpyxl.styles.Alignment(
                horizontal='center', vertical='bottom')

            sheet['B12'] = int(sheet['B10'].value) * int(sheet['B11'].value)
            sheet['B12'].font = Font(name='Century', size=12)
            sheet['B12'].alignment = openpyxl.styles.Alignment(
                horizontal='center', vertical='bottom')

            sheet['C12'] = int(sheet['C10'].value) * int(sheet['C11'].value)
            sheet['C12'].font = Font(name='Century', size=12)
            sheet['C12'].alignment = openpyxl.styles.Alignment(
                horizontal='center', vertical='bottom')

            sheet['E12'] = str(sheet['B12'].value + sheet['C12'].value)+'/140'
            sheet['E12'].font = Font(name='Century', size=12, color='0000FF')
            sheet['E12'].alignment = Alignment(
                horizontal='center', vertical='bottom')

            wb.save(os.path.join(os.getcwd(), "sample_output",
                    "marksheet", (row[6].upper()+".xlsx")))


app = Flask(__name__)

@app.route('/')
def upload_file():
    return render_template('homepage.html')


@app.route("/", methods=['POST'])
def uploader_file():
    path1 = request.files['master_roll']
    path2 = request.files['response']
    process = request.form['process']
    print(process)
    if path1.filename != '':
        path1 = os.path.join('upload', 'master_roll.csv')

    if path2.filename != '':
        path2 = os.path.join('upload', 'response.csv')
    positive = request.form["positive"]
    negative = request.form["negative"]

    if process == "genrolmarksheet":
        generate_marksheet(path2, positive, negative)
    
    if process == "conmarksheet":
        concise(path2, positive, negative)

    if process == "email":
        send_email(path2)

    return render_template('message.html')


if __name__ == "__main__":
    app.run(debug=True)
