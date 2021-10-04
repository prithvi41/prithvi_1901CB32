import csv
import openpyxl 

ltp_dict=dict() #dictionary to store subcode-ltp
sub_name=dict() #dictionary to store subcode-subname
with open("subjects_master.csv","r") as file:
     reader=csv.DictReader(file)
     for row in reader:
         ltp_dict[row["subno"]]=row["ltp"]
         sub_name[row["subno"]]=row["subname"]
 
name_dict=dict() #dictionary to store rollno-name
with open("names-roll.csv","r") as file:
      reader=csv.DictReader(file)
      for row in reader:
          name_dict[row["Roll"]]=row["Name"]


def generate_marksheet(name_dict, sub_name, ltp_dict, data_of_a_rollno,  int_grading, roll_no):
    wb = openpyxl.Workbook() 
    sheet = wb.active 
    sheet.title="Overall"  #title of first index sheet

    #common header in all sem data
    all_reqdata=['Sl No.','Subject No.','Subject Name','L-T-P','Credit','Subject Type','grade']
    curr_sem=0  #current sem set to 1 initially
    sheet_index=1  #current index set to 1 initially
    sl_no=1

    for row in data_of_a_rollno: 
        sheet_title="Sem"+row["Sem"]     
        if int(row["Sem"])!=curr_sem: 
            sl_no=1                  #again make  serial number as 1 
            curr_sem=int(row["Sem"])   #updating curr_sem to next semester
            wb.create_sheet(index=sheet_index, title=sheet_title) 
            wb[sheet_title].append(all_reqdata)         #add same heading(all_reqdata) to the new sheet 
            sheet_index=sheet_index+1           
        
        #create a list named grade with required student_info and append it to the sheet 
        sem_data=[sl_no, row["SubCode"], sub_name[row["SubCode"]], ltp_dict[row["SubCode"]], row["Credit"], row["Sub_Type"], row["Grade"]] 
        wb[sheet_title].append(sem_data) 
        sl_no=sl_no+1                

   #giving name of req cell in "overall" sheet and writing req data
    sheet['A1']="Roll No."           
    sheet['B1']=roll_no                
    sheet['A2']="Name of Student"      
    sheet.cell(row = 2, column = 2).value = name_dict[roll_no]   
    sheet['A3']="Discipline"         
    Discipline=roll_no[4:6]     #name of branch by slicing roll no at req places 
    sheet['B3']=Discipline           
    sheet['A4']="Semester No."       
    sheet['A5']="Semester wise credit Taken"     
    sheet['A6']="SPI"                            
    sheet['A7']="Total Credits Taken"            
    sheet['A8']="CPI" 

    sheets_list=wb.sheetnames     # all the sheets listed
    credit_taken=0          # total credit_of_a_course taken by student                 
    credit_obtained=0   

    for i in range(1,len(sheets_list)): 
        sheet_for_sem=wb[sheets_list[i]]     #sem wise sheet 
        sem_no=sheets_list[i][3:]    # slicing sem no

        column_letter=openpyxl.utils.get_column_letter(i+1)  
        sheet[column_letter+str(4)]=int(sem_no)             

        semwise_credit=0                      # total credit per semester taken  
        semwise_obtained=0                     
        row_count = sheet_for_sem.max_row         #total num

        for j in range(2, row_count + 1):  
            #credit_of_a_course and grade of a particular course
            credit_of_a_course=int(sheet_for_sem['E'+str(j)].value)  
            grade_of_a_course=sheet_for_sem['G'+str(j)].value 

            semwise_credit=semwise_credit+credit_of_a_course     # total credit_of_a_course in a given semester 
            semwise_obtained=semwise_obtained+(credit_of_a_course*int_grading[grade_of_a_course])  

        credit_taken=credit_taken+semwise_credit     #total credit till  given semester 
        credit_obtained=credit_obtained+semwise_obtained
        SPI=semwise_obtained/semwise_credit     # SPI using formula 
        SPI=round(SPI,2)    
        CPI=credit_obtained/credit_taken        # CPI using formula 
        CPI=round(CPI,2) 
        #now write all value in specific place                    
        sheet[column_letter+str(5)]=semwise_credit   
        sheet[column_letter+str(6)]=SPI               
        sheet[column_letter+str(7)]=credit_taken      
        sheet[column_letter+str(8)]=CPI             
     
    wb.save(r'output\\'+'%s.xlsx'%roll_no)
    return


int_grading={'AA':10, 'AB':9, 'BB':8, 'BC':7, 'CC':6, 'CD':5, 'DD':4, 'DD*':4, 'F*':0, 'F':0, 'I':0, ' BB':8} #dict to map grade 

roll_no="0401CS01" #first roll number
data_of_a_rollno =[]     #list to store all the data from grades.csv 
with open("grades.csv","r") as file:
    reader=csv.DictReader(file)          
    for row in reader:     
       if row["Roll"]==roll_no:         
         data_of_a_rollno.append(row)
       else: 
        # calling fxn for a particular roll and then storing next rollno
         generate_marksheet(name_dict, sub_name, ltp_dict, data_of_a_rollno,  int_grading, roll_no) 
         data_of_a_rollno=[]     
         data_of_a_rollno.append(row) 
         roll_no=row["Roll"] 

#calling again for last leftout rollno
generate_marksheet(name_dict, sub_name, ltp_dict, data_of_a_rollno,  int_grading, roll_no)





